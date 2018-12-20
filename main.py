import os
from openpyxl import load_workbook
from flashcard import *
from exportExcel import export_to_excel

# set input path
os.chdir(r"/home/skibojq/projects/flashcards-anki/files")
wb = load_workbook('./input1.xlsx')
ws = wb.active

# create dictionary with words
data = {}
for cell in ws['A:A']:
    print(cell.value)
    try:
        data['%s' % cell.value] = CreateFlashcard(cell.value)
    except:
        pass

export_to_excel(data)

print("finish")
